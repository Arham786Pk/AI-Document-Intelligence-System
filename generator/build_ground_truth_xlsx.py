"""Task 2 — build ground_truth.xlsx from ground_truth.csv.

The CSV is the source of truth (easy to diff in git). This script wraps
it in an Excel workbook with proper formatting for human review.
"""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
CSV_PATH = ROOT / "docs" / "ground_truth.csv"
XLSX_PATH = ROOT / "docs" / "ground_truth.xlsx"

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")

COL_WIDTHS = {
    "document_name": 46,
    "doc_type": 18,
    "language": 10,
    "modality": 12,
    "source": 18,
    "project_id": 14,
    "supplier": 32,
    "material": 30,
    "quantity": 14,
    "date": 12,
    "notes": 60,
}


def main() -> None:
    rows = list(csv.DictReader(CSV_PATH.open(encoding="utf-8")))
    headers = list(rows[0].keys())

    wb = Workbook()
    ws = wb.active
    ws.title = "Ground Truth"

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        ws.append([row[h] for h in headers])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP

    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = COL_WIDTHS.get(h, 20)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(XLSX_PATH)
    print(f"wrote {XLSX_PATH.relative_to(ROOT)} ({len(rows)} rows)")


if __name__ == "__main__":
    main()
